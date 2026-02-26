from rest_framework import status
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import permission_classes, action
from rest_framework.permissions import IsAdminUser, IsAuthenticated


import os
import io
import csv
from django.http import HttpResponse
from django.conf import settings
from django.shortcuts import render
from django.db.models import Sum, Q, Count
from django.db import transaction 
from datetime import datetime, timedelta
from django.db.models.functions import Coalesce
from django.db.models import DecimalField, ExpressionWrapper, F, Value, Case, When

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import chardet


from .serializers import VentasSerializers, DetalleVentasSerializers
from .models import VentasDetallada
from pedidos.models import Pedidos

from core.permissions import TienePermisoDinamico

# ===== ESTILOS =====
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4B5563")  # gris oscuro
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")  # gris claro

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

NUM_FMT = '#,##0.00'  # miles + 2 decimales



# Create your views here.

class VentasViewSet(viewsets.ViewSet):

    def get_serializer_class(self):
        if self.action in ['detalle', 'detail_category']: 
            return DetalleVentasSerializers
        return VentasSerializers
    
    def get_permissions(self):
        """
        Mapeo de acciones a permisos específicos
        """
        # 1. ACCIÓN: Sincronizar con API externa (POST)
        if self.action == 'create':
            return [IsAuthenticated(), TienePermisoDinamico('cargar-datos')]

        # 2. ACCIÓN: Ver listado para el combo/dropdown (GET)
        if self.action == 'detalle':
            return [IsAuthenticated(), TienePermisoDinamico('ver_detalle_ventas')]

        # 3. ACCIÓN: Exportar Excel (GET)
        if self.action == 'exportar_ventas':
            return [IsAuthenticated(), TienePermisoDinamico('ver_ventas')]
        
        # 4. ACCIÓN: resumen (GET)
        if self.action == 'resumen':
            return [IsAuthenticated(), TienePermisoDinamico('ver_ventas')]

        # Por defecto, cualquier otra cosa requiere estar logueado
        return [IsAdminUser()]

    def get_queryset_agrupado(self, inic, fin, grupo, orden):
        # Usar el modelo directamente desde el serializer
        model = self.get_serializer_class().Meta.models
        return model.objects.filter(Fecha_Comprobante__range=(inic, fin)) \
            .values(*grupo) \
            .annotate(
                suma_subtotal_neto=Sum('Subtotal_Neto'),
                suma_subtotal_final=Sum('Subtotal_Final'),
                suma_costo_neto=Sum('Costo_Neto')
            ) \
            .order_by()
    
    
    @action(detail=False, methods=['get'])
    def detalle(self, request):
        tipo = request.query_params.get("tipo")
        codigo = request.query_params.get("codigo")
        inicio = request.query_params.get("inicio")
        fin = request.query_params.get("fin")

        # -----------------------------
        # 1. Validaciones
        # -----------------------------
        if not tipo or not inicio or not fin:
            return Response(
                {"error": "Parámetros requeridos: tipo, inicio, fin"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            inic = datetime.strptime(inicio, "%Y-%m-%d").date()
            fini = datetime.strptime(fin, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Formato de fecha inválido. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------
        # 2. Query base
        # -----------------------------
        qs = (
            VentasDetallada.objects
            .filter(Fecha_Comprobante__range=(inicio, fin))
            .select_related()   # si luego agregás FK
            
        )
        

       # print(codigo, tipo, inicio, fin)
    

        # -----------------------------
        # 3. Filtro por tipo
        # -----------------------------
        if tipo == "articulo" and codigo:
            #print("Pedido de articulo")
            qs = qs.filter(Codigo_Articulo=codigo)

        elif tipo == "cliente" and codigo:
            qs = qs.filter(Cliente=codigo)

        elif tipo == "canal" and codigo:
            qs = qs.filter(IdCanal=codigo)

        elif tipo == "vendedora" and codigo:
            qs = qs.filter(Vendedor=codigo)

        elif tipo == "localidad" and codigo:
            qs = qs.filter(Localidad=codigo)

        # -----------------------------
        # 4. Construcción del detalle
        # -----------------------------
    
        data = qs.annotate(
                IVA_calc=ExpressionWrapper(
                    F("Subtotal_Final") - F("Subtotal_Neto"),
                    output_field=DecimalField(max_digits=15, decimal_places=2)
                ),

                Utilidad=ExpressionWrapper(
                    F("Subtotal_Final") - F("Costo_Neto"),
                    output_field=DecimalField(max_digits=15, decimal_places=2)
                ),

                Cmg=Case(
                    When(
                        Subtotal_Final__gt=0,
                        then=ExpressionWrapper(
                            (F("Subtotal_Final") - F("Costo_Neto")) * Value(100.0) / F("Subtotal_Final"),
                            output_field=DecimalField(max_digits=7, decimal_places=2)
                        )
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=7, decimal_places=2)
                )
            ).values(
                "Comprobante",
                "Descripcion_Comprobante",
                "Letra",
                "Serie",
                "Numero",
                "Informado",
                "Motivo_Devolucion",
                "Descripcion_Motivo_Devolucion",
                "Fecha_Comprobante",
                "Sucursal",
                "Descripcion_Sucursal",
                "Vendedor",
                "Descripcion_Vendedor",
                "Numero_Pedido",
                "Cliente",
                "Razon_Social",
                "Zona",
                "Localidad",
                "Nro_Linea",
                "Codigo_Articulo",
                "Descripcion_Articulo",
                "Unidades_Bulto",
                "Categorias",
                "Descripcion_Categorias",
                "Familia",
                "Descripcion_Familia", 
                "Costo_Neto",
                "Bultos_Total",
                "Subtotal_Neto",
                "IVA_calc",
                "Subtotal_Final",
                "Utilidad",
                "Cmg",    
                "IdCanal",
                "Canal"
            )
        #print("Total Registro", len(data))
        #serializer = DetalleVentasSerializers(data, many=True)
        data2=pd.DataFrame(data)
        columnas_numericas = [
            'Subtotal_Neto', 'Subtotal_Final',
            'Costo_Neto', 'Bultos_Total'
        ]
        data2[columnas_numericas] = data2[columnas_numericas].apply(
            pd.to_numeric, errors='coerce'
        ).fillna(0)
        #data2.to_excel("VentasDetalla.xlsx", sheet_name='Ventas', index=False)

        totales = qs.aggregate(
            costo_neto=Sum('Costo_Neto'),
            subtotal_neto=Sum('Subtotal_Neto'),
            subtotal_final=Sum('Subtotal_Final'),
            # Calculamos IVA y Utilidad sobre los totales para mayor precisión
            iva_calc=Sum(F('Subtotal_Final') - F('Subtotal_Neto')),
            utilidad=Sum(F('Subtotal_Final') - F('Costo_Neto'))
        )

        # Cálculo manual del CMG General (Margen de Contribución)
        if totales['subtotal_final'] and totales['subtotal_final'] > 0:
            totales['cmg'] = (totales['utilidad'] * 100) / totales['subtotal_final']
        else:
            totales['cmg'] = 0

        # Importante: Cambiamos la respuesta para enviar un objeto con 'items' y 'totalesGenerales'
        respuesta = {
            "items": list(data),
            "totalesGenerales": totales
        }
        # -----------------------------
        # 5. Respuesta
        # -----------------------------
        return Response(respuesta, status=status.HTTP_200_BAD_REQUEST if not data else status.HTTP_200_OK)
    
        #return Response({"tipo": tipo,"codigo": codigo, "inicio": inic, "fin": fin, "items": list(data) }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def exportar_ventas(self, request):
        
        data = self.list(request).data
        ventas = data["Ventas"]
        fileName = "reporte_ventas(" + data['rango']['inicio'].strftime('%d-%m-%Y') + " - " + data['rango']['fin'].strftime('%d-%m-%Y') + ").xlsx"

        wb = Workbook()
        wb.remove(wb.active)  # borrar hoja default

        # Configuración por grupo
        config = {
            "Canal": ("Canal", 22),
            "Vendedoras": ("Descripcion_Vendedor", 35),
            "Localidad": ("Localidad", 35),
            "Articulos": ("Descripcion_Articulos", 50),
            "Clientes": ("Razon_Social", 55),
            "Fecha": ("Fecha_Comprobante", 20),
        }

        TOP_N = 10
        GRUPOS_TOP = {"Articulos", "Clientes", "Localidad"}

        for grupo, filas in ventas.items():
            if grupo not in config:
                continue  # Salta claves vacías o no configuradas

            key_field, ancho_col_a = config[grupo]

            ws = wb.create_sheet(grupo)

            # Margen para la gráfica
            for _ in range(18):
                ws.append([])

            ws.append([
                f"Inicio: {data['rango']['inicio'].strftime('%d-%m-%Y')}",
                f"Fin: {data['rango']['fin'].strftime('%d-%m-%Y')}"
            ]) 
            # Orden especial FECHA
            if grupo == "Fecha":
                filas = sorted(filas, key=lambda x: x["Fecha_Comprobante"])

            elif grupo in GRUPOS_TOP:
                # Subtotal Final DESC
                filas = sorted(
                    filas,
                    key=lambda x: x.get("Subtotal_Final", 0),
                    reverse=True
                )

            header_row, first_data, last_data, total_row = (
                self.escribir_tabla_con_totales(ws, filas, key_field)
            )

            # Formatos
            self.formatear_header(ws, header_row)
            self.formatear_numeros(ws, first_data, last_data)
            self.formatear_totales(ws, total_row)
            self.ajustar_columnas(ws, ancho_col_a)

            # 🔹 Gráfica (solo TOP N si aplica)
            self.crear_grafica_arriba(
                ws,
                first_data,
                last_data,
                (f"Resumen de las TOP {TOP_N} de Ventas por {grupo}" if grupo in GRUPOS_TOP else f"Resumen de Ventas por {grupo}"),
                solo_top_n=(grupo in GRUPOS_TOP),
                top_n=TOP_N
            )

            # Lógica especial FECHA
            if grupo == "Fecha":
                ws.append([])
                self.agregar_acumulado_fecha(ws, last_data)
        print(fileName)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{fileName}"'
        wb.save(response)

        return response

    @action(detail=False, methods=['get'])
    def resumen(self, request):
        inicio = request.query_params.get("inicio")
        final = request.query_params.get("fin")
        if not inicio or not final:
            return Response({'error': 'Faltan parámetros de fecha (inicio, fin).'}, status=400)
        try:
            inic = datetime.strptime(inicio, "%Y-%m-%d").date()
            fin = datetime.strptime(final, "%Y-%m-%d").date()
            if fin.weekday() == 6:  # Si es domingo
                fin -= timedelta(days=1)
            fin2 = fin - timedelta(days=1)
            # Si el resultado es domingo (weekday es 6), restar otro día
            if fin2.weekday() == 6 and fin2.day > 1 :
                fin2 -= timedelta(days=1)
        except ValueError:
            return Response({'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)
        qsacum = self.get_queryset_agrupado(inic, fin2, ['Comprobante'], [])
        qs = self.get_queryset_agrupado(inic, fin, ['Comprobante'], [])
        if len(qsacum) == 0:
            #return Response({'error': 'No hay datos acumulados hasta la fecha anterior al rango proporcionado.'}, status=404)
            qsacum = [
                    {"Comprobante": "FCVTA", "suma_subtotal_final": 0},
                    {"Comprobante": "DVVTA",  "suma_subtotal_final": 0}
                ]
        if len(qs) == 0:
            return Response({'error': 'No hay datos en el rango de fechas proporcionado.'}, status=404)
        else:
            if True: #Ventas Resumidas
                acum = qsacum[0]['suma_subtotal_final'] + qsacum[1]['suma_subtotal_final']
                Total = qs[0]['suma_subtotal_final'] + qs[1]['suma_subtotal_final'] 
                Ventas = {
                            "totalV": Total, 
                            "totalF": qs[1]['suma_subtotal_final'], 
                            "totalNC": qs[0]['suma_subtotal_final'],
                            "acum": acum,
                            "acumF": qsacum[1]['suma_subtotal_final'],
                            "acumNC": qsacum[0]['suma_subtotal_final'],
                            "dif": Total - acum,
                            "final": fin.strftime('%d-%m-%Y'),
                            "final2": fin2.strftime ('%d-%m-%Y')
                        }
            if True: #Pedidos Resumidos
                    stats = Pedidos.objects.filter(fecalta__range=(inic, fin)).aggregate(
                    # Conteos
                    total_cant=Count('id'),
                    pf_cant=Count('id', filter=Q(facturado=True)),
                    pp_cant=Count('id', filter=Q(anulado=False, facturado=False)),
                    pc_cant=Count('id', filter=Q(anulado=True)),
                    
                    # Sumas (ajusta 'monto' al nombre real de tu campo, ej: 'total_pedido')
                    total_monto=Sum('total'),
                    pf_monto=Sum('total', filter=Q(facturado=True)),
                    pp_monto=Sum('total', filter=Q(anulado=False, facturado=False)),
                    pc_monto=Sum('total', filter=Q(anulado=True))
                )


                    R_Pedidos = {
                        "total": { "cantidad": stats['total_cant'], "monto": stats['total_monto'] or 0 },
                        "totalPF": { "cantidad": stats['pf_cant'], "monto": stats['pf_monto'] or 0 },
                        "totalPP": { "cantidad": stats['pp_cant'], "monto": stats['pp_monto'] or 0 },
                        "totalPC": { "cantidad": stats['pc_cant'], "monto": stats['pc_monto'] or 0 } 
                        }

            resumen = {"Ventas": Ventas, "Pedidos": R_Pedidos}   
        return Response({
            "message": "Resumen generado con éxito",
            "rango": {"inicio": inic, "fin": fin},
            "Resumen": resumen
        }, status=status.HTTP_200_OK)

    def list(self, request):
        inicio = request.query_params.get("inicio")
        final = request.query_params.get("fin")

        if not inicio or not final:
            return Response({'error': 'Faltan parámetros de fecha (inicio, fin).'}, status=400)

        try:
            inic = datetime.strptime(inicio, "%Y-%m-%d").date()
            fin = datetime.strptime(final, "%Y-%m-%d").date()
        except ValueError:
            return Response({'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)

        # Definición de las agrupaciones para iterar (DRY - Don't Repeat Yourself)
        configuraciones = {
            'Canal': (['Canal', 'Comprobante', 'IdCanal'], ['Canal'], 'Canal'),
            'Vendedoras': (['Vendedor', 'Descripcion_Vendedor', 'Comprobante'], ['Vendedor'], 'Vendedor'),
            'Localidad': (['Zona', 'Localidad', 'Comprobante'], ['Zona'], 'Localidad'),
            'Articulos': (['Codigo_Articulo', 'Descripcion_Articulo', 'Comprobante'], ['Codigo_Articulo'], 'Codigo_Articulo'),
            'Clientes': (['Cliente', 'Razon_Social', 'Comprobante'], ['Cliente'], 'Cliente'),
            'Fecha': (['Fecha_Comprobante', 'Comprobante'], ['Fecha_Comprobante'], 'Fecha_Comprobante'),
            '': (['Comprobante'], [], 'Comprobante'),
        }

        resultados = {}
        for clave, (grupo, orden, proc_key) in configuraciones.items():
            qs = self.get_queryset_agrupado(inic, fin, grupo, orden)
        
            if clave == 'Canal':
                print("inic ", inic, "fin ", fin, " ")
            #for row in qs:
            #     print(f"DEBUG ROW: Canal={row['Canal']}, IdCanal={row['IdCanal']}, Suma={row['suma_subtotal_neto']}")
            resultados[clave] = self.procesar_seccion(qs, proc_key)

        return Response({
            "message": "Reporte generado con éxito",
            "rango": {"inicio": inic, "fin": fin},
            "Ventas": resultados
        }, status=status.HTTP_200_OK)
   
    def create(self, request):

        # -----------------------------------------
        # 0. Validación del archivo recibido
        # -----------------------------------------
        
        archivo = request.FILES.get('file')
        if not archivo:
            return Response(
                {'error': 'Debe enviar un archivo en el campo "file".'},
                status=status.HTTP_400_BAD_REQUEST
            )

        nombre = archivo.name.lower()

        # Validar extensión
        extensiones_permitidas = ('.xls', '.xlsx', '.txt', '.csv')
        if not nombre.endswith(extensiones_permitidas):
            return Response(
                {'error': 'Formato no permitido. Debe ser XLS, XLSX, TXT o CSV.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tamaño (20 MB)
        if archivo.size > 20 * 1024 * 1024:
            return Response(
                {'error': 'El archivo supera el tamaño máximo permitido (20 MB).'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------------------
        # 1. Columnas requeridas
        # -----------------------------------------
        columnas_requeridas = [
            'Comprobante', 'Descripcion Comprobante', 'Letra',
            'Serie \\ Punto de venta', 'Numero', 'Informado',
            'Motivo Rechazo / Devolucion', 'Descripcion Motivo Rechazo / Devolucion',
            'Fecha Comprobante', 'Sucursal', 'Descripcion Sucursal',
            'Vendedor', 'Descripcion Vendedor', 'Numero de Pedido',
            'Cliente', 'Razon Social', 'Cod. Postal', 'Localidad',
            'Nro. de Linea', 'Codigo de Articulo', 'Descripcion de Articulo',
            'Unidades por Bulto', 'CATEGORIAS', 'Descripción CATEGORIAS',
            'FAMILIA', 'Descripción FAMILIA', 'Precio de compra Neto',
            'Bultos Total', 'Subtotal Neto', 'Subtotal Final'
        ]

        dfs = []
        comprobantes_validos = ['FCVTA', 'DVVTA']
        clientes_no_validos = [
            'ABANS CONSTRUCCION S.R.L', 'ABANS PROCESOS SRL', 'BARTRADE SRL',
            'MAIN SUPPORT S.R.L.', 'MAS ACTIVOS S. A', 'SACEM S A', 'SINDAL AR SA', 'OPIFEX PRO S.R.L',
            'LEVEL TRUCK', 'NIVEL TRUCK SRL', 'FULL BENEFITS SA', 'APAHIE S.R.L'
        ]

        # -----------------------------------------
        # 2. Lectura del archivo
        # -----------------------------------------

        # SOLO .xlsx es Excel real
        if nombre.endswith('.xlsx'):
            try:
                df_excel = pd.read_excel(
                    archivo,
                    usecols=columnas_requeridas,
                    engine='openpyxl'
                )

                mask = (
                    df_excel['Comprobante'].isin(comprobantes_validos) &
                    ~df_excel['Razon Social'].isin(clientes_no_validos)
                )

                dfs.append(df_excel[mask].copy())

            except Exception as e:
                return Response(
                    {'error': f'Error al leer el archivo Excel: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # TODO LO DEMÁS → CSV/TXT (incluye .xls disfrazado)
        else:
            try:
                # Detectar encoding con muestra
                raw_sample = archivo.read(50000)
                det = chardet.detect(raw_sample)
                encoding_detectado = det['encoding'] or 'latin1'

                archivo.seek(0)
                raw = archivo.read()
                texto = raw.decode(encoding_detectado, errors='replace')

                # Detectar separador
                try:
                    sniffer = csv.Sniffer()
                    dialect = sniffer.sniff(texto[:2000])
                    sep = dialect.delimiter
                except Exception:
                    sep = '\t'

                archivo_io = io.StringIO(texto)

                chunks = pd.read_csv(
                    archivo_io,
                    sep=sep,
                    usecols=columnas_requeridas,
                    chunksize=30000,
                    low_memory=False
                )

                for chunk in chunks:
                    mask = (
                        chunk['Comprobante'].isin(comprobantes_validos) &
                        ~chunk['Razon Social'].isin(clientes_no_validos)
                    )
                    dfs.append(chunk[mask].copy())

            except Exception as e:
                return Response(
                    {'error': f'Error al leer el archivo CSV/TXT: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # -----------------------------------------
        # 3. Unificación
        # -----------------------------------------
        if not dfs:
            return Response({'error': 'No hay datos válidos'}, status=status.HTTP_400_BAD_REQUEST)

        df = pd.concat(dfs, ignore_index=True)
        del dfs

        # -----------------------------------------
        # 4. Sanitización de datos
        # -----------------------------------------

        # Normalizar Informado
        df['Informado'] = df['Informado'].astype(str).str.strip().str.upper()

        # Campos enteros
        cols_int = [
            'Vendedor', 'Numero', 'Motivo Rechazo / Devolucion', 'Sucursal',
            'Numero de Pedido', 'Cliente', 'Cod. Postal', 'Nro. de Linea',
            'Codigo de Articulo', 'Unidades por Bulto', 'CATEGORIAS', 'FAMILIA'
        ]
        for col in cols_int:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

        # Campos decimales
        cols_decimal = [
            'Precio de compra Neto', 'Bultos Total', 'Subtotal Neto', 'Subtotal Final'
        ]
        for col in cols_decimal:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        # Cálculos adicionales
        df['Costo_Neto'] = df['Precio de compra Neto'] * 1.21 * df['Bultos Total']
        df['Subtotal Neto'] = np.where(
            df['Informado'] == 'NO',
            df['Subtotal Final'],
            df['Subtotal Neto']
        )

        # -----------------------------------------
        # 5. Canal e IdCanal
        # -----------------------------------------
        mapa_canal = {
            'Vendedoras': [1, 3, 4, 5, 6, 11, 14, 16, 17, 18],
            'Local': [0,7],
            'Mayoristas': [9]
        }

        vendedor_to_canal = {}
        for canal, lista in mapa_canal.items():
            for v in lista:
                vendedor_to_canal[v] = canal

        df['Canal'] = df['Vendedor'].map(vendedor_to_canal).fillna('Otros')

        orden = {canal: i + 1 for i, canal in enumerate(mapa_canal.keys())}
        df['IdCanal'] = df['Canal'].map(orden).fillna(4).astype(int)

        # -----------------------------------------
        # 6. Fechas y strings
        # -----------------------------------------
        df['Fecha Comprobante'] = pd.to_datetime(
            df['Fecha Comprobante'],
            errors='coerce',
            dayfirst=True
        ).dt.date

        df['Descripcion Vendedor'] = df['Descripcion Vendedor'].fillna('Directo')
        df['Descripcion Motivo Rechazo / Devolucion'] = df[
            'Descripcion Motivo Rechazo / Devolucion'
        ].fillna('')

        # -----------------------------------------
        # 7. Mapeo final de columnas al modelo
        # -----------------------------------------
        mapeo = {
            'Descripcion Comprobante': 'Descripcion_Comprobante',
            'Serie \\ Punto de venta': 'Serie',
            'Motivo Rechazo / Devolucion': 'Motivo_Devolucion',
            'Descripcion Motivo Rechazo / Devolucion': 'Descripcion_Motivo_Devolucion',
            'Fecha Comprobante': 'Fecha_Comprobante',
            'Descripcion Sucursal': 'Descripcion_Sucursal',
            'Descripcion Vendedor': 'Descripcion_Vendedor',
            'Numero de Pedido': 'Numero_Pedido',
            'Razon Social': 'Razon_Social',
            'Cod. Postal': 'Zona',
            'Nro. de Linea': 'Nro_Linea',
            'Codigo de Articulo': 'Codigo_Articulo',
            'Descripcion de Articulo': 'Descripcion_Articulo',
            'Unidades por Bulto': 'Unidades_Bulto',
            'CATEGORIAS': 'Categorias',
            'Descripción CATEGORIAS': 'Descripcion_Categorias',
            'FAMILIA': 'Familia',
            'Descripción FAMILIA': 'Descripcion_Familia',
            'Bultos Total': 'Bultos_Total',
            'Subtotal Neto': 'Subtotal_Neto',
            'Subtotal Final': 'Subtotal_Final'
        }
        df.rename(columns=mapeo, inplace=True)

        columnas_validas = [
            f.name for f in VentasDetallada._meta.fields if f.name != 'id'
        ]
        df_final = df.reindex(columns=columnas_validas)

        # -----------------------------------------
        # 8. Bulk create con manejo de conflictos
        # -----------------------------------------
        columnas_pk = ['Comprobante', 'Letra', 'Serie', 'Numero', 'Nro_Linea', 'Codigo_Articulo']
        update_fields = [c for c in columnas_validas if c not in columnas_pk]
        registros = df_final.to_dict('records')
        instancias = [VentasDetallada(**row) for row in registros]

        with transaction.atomic():
            VentasDetallada.objects.bulk_create(
                instancias,
                batch_size=5000,
                update_conflicts=True,
                unique_fields=columnas_pk,
                update_fields=update_fields
            )

        return Response(
            {'message': f'Éxito: {len(instancias)} registros procesados.'},
            status=status.HTTP_200_OK
        )

    def procesar_seccion(self, queryset, key):
        resultado = {}

        # ============================================================
        # 1. Construcción de FCVTA / DVVTA y recolección de metadatos
        # ============================================================
        for item in queryset:
            grupo = item[key]
            comp = item["Comprobante"]

            if grupo not in resultado:
                resultado[grupo] = {
                    "FCVTA": {"Costo": 0, "Neto": 0, "Final": 0},
                    "DVVTA": {"Costo": 0, "Neto": 0, "Final": 0},
                    "meta": {}
                }

            # Acumulación de totales (Mantenemos el signo que viene de la DB)
            resultado[grupo][comp]["Costo"] += item["suma_costo_neto"]
            resultado[grupo][comp]["Neto"] += item["suma_subtotal_neto"]
            resultado[grupo][comp]["Final"] += item["suma_subtotal_final"]

            # Guardar metadatos (Solo si existen en el item)
            claves_meta = ["IdCanal", "Descripcion_Vendedor", "Zona", "Descripcion_Articulo", "Razon_Social"]
            for c in claves_meta:
                if c in item:
                    resultado[grupo]["meta"][c] = item[c]

        # ============================================================
        # 2. Construcción del JSON final con datos consolidados
        # ============================================================
        salida = []

        for grupo, datos in resultado.items():
            # IVA por tipo (Diferencia entre Final y Neto)
            iva_f = datos["FCVTA"]["Final"] - datos["FCVTA"]["Neto"]
            iva_d = datos["DVVTA"]["Final"] - datos["DVVTA"]["Neto"]

            # CONSOLIDACIÓN: Se usa SUMA (+) porque DVVTA ya es negativo en los datos
            costo = datos["FCVTA"]["Costo"] + datos["DVVTA"]["Costo"]
            neto = datos["FCVTA"]["Neto"] + datos["DVVTA"]["Neto"]
            iva = iva_f + iva_d
            final = datos["FCVTA"]["Final"] + datos["DVVTA"]["Final"]
            
            utilidad = final - costo
            cmg = (utilidad / final * 100) if final != 0 else 0

            fila = {
                key: grupo,
                "Costo": costo,
                "Subtotal_Neto": neto,
                "IVA": iva,
                "Subtotal_Final": final,
                "Utilidad": utilidad,
                "Cmg": cmg,
                "Detalle": {
                    "FCVTA": {
                        "Costo": datos["FCVTA"]["Costo"],
                        "Subtotal_Neto": datos["FCVTA"]["Neto"],
                        "IVA": iva_f,
                        "Subtotal_Final": datos["FCVTA"]["Final"],
                        "Utilidad": datos["FCVTA"]["Final"] - datos["FCVTA"]["Costo"],
                        "Cmg": ((datos["FCVTA"]["Final"] - datos["FCVTA"]["Costo"]) / datos["FCVTA"]["Final"] * 100)
                            if datos["FCVTA"]["Final"] != 0 else 0
                    },
                    "DVVTA": {
                        "Costo": datos["DVVTA"]["Costo"],
                        "Subtotal_Neto": datos["DVVTA"]["Neto"],
                        "IVA": iva_d,
                        "Subtotal_Final": datos["DVVTA"]["Final"],
                        "Utilidad": datos["DVVTA"]["Final"] - datos["DVVTA"]["Costo"],
                        "Cmg": ((datos["DVVTA"]["Final"] - datos["DVVTA"]["Costo"]) / datos["DVVTA"]["Final"] * 100)
                            if datos["DVVTA"]["Final"] != 0 else 0
                    }
                }
            }

            # Inyectar metadatos al nivel principal de la fila
            meta = datos["meta"]
            if "IdCanal" in meta: fila["IdCanal"] = meta["IdCanal"]
            if "Descripcion_Vendedor" in meta: fila["Descripcion_Vendedor"] = meta["Descripcion_Vendedor"]
            if "Zona" in meta: fila["Zona"] = meta["Zona"]
            if "Descripcion_Articulo" in meta: fila["Descripcion_Articulos"] = meta["Descripcion_Articulo"]
            if "Razon_Social" in meta: fila["Razon_Social"] = meta["Razon_Social"]

            salida.append(fila)

        # ============================================================
        # 3. ORDENAR COLUMNAS SEGÚN SECCIÓN
        # ============================================================
        if key == "Canal":
            salida = [{"IdCanal": f.get("IdCanal", 0), "Canal": f["Canal"], **{k: v for k, v in f.items() if k not in ["IdCanal", "Canal"]}} for f in salida]
        elif key == "Vendedor":
            salida = [{"Vendedor": f["Vendedor"], "Descripcion_Vendedor": f.get("Descripcion_Vendedor", ""), **{k: v for k, v in f.items() if k not in ["Vendedor", "Descripcion_Vendedor"]}} for f in salida]
        elif key == "Localidad":
            salida = [{"Zona": f.get("Zona", ""), "Localidad": f["Localidad"], **{k: v for k, v in f.items() if k not in ["Zona", "Localidad"]}} for f in salida]
        elif key == "Codigo_Articulo":
            salida = [{"Articulo": f["Codigo_Articulo"], "Descripcion_Articulos": f.get("Descripcion_Articulos", ""), **{k: v for k, v in f.items() if k not in ["Codigo_Articulo", "Descripcion_Articulos"]}} for f in salida]
        elif key == "Cliente":
            salida = [{"Cliente": f["Cliente"], "Razon_Social": f.get("Razon_Social",""), **{k: v for k, v in f.items() if k not in ["Cliente", "Razon_Social"]}} for f in salida]

        return salida

    def escribir_tabla_con_totales(self, ws, filas, key_field):
       
        ws.append([
            key_field,
            #"Costo",
            # "Subtotal Neto",
            # "IVA",
            "Subtotal Final",
            # "Utilidad",
            # "CMG %"
        ])
        header_row = ws.max_row

        ws.auto_filter.ref = f"A{header_row}:B{header_row}"
        for f in filas:
            ws.append([
                f.get(key_field),
                # f["Costo"],
                # f["Subtotal_Neto"],
                # f["IVA"],
                f["Subtotal_Final"],
                # f["Utilidad"],
                # f["Cmg"]
            ])

        first_data = header_row + 1
        last_data = ws.max_row

        # TOTAL
        ws.append([
            "TOTAL",
            f"=SUM(B{first_data}:B{last_data})",
            #f"=SUM(C{first_data}:C{last_data})",
            #f"=SUM(D{first_data}:D{last_data})",
            #f"=SUM(E{first_data}:E{last_data})",
            #f"=SUM(F{first_data}:F{last_data})",
            #""
        ])
        total_row = ws.max_row

        return header_row, first_data, last_data, total_row

    def crear_grafica_arriba(
        self,
        ws,
        first_data,
        last_data,
        titulo,
        solo_top_n=False,
        top_n=10
    ):

        # Determinar rango de la gráfica
        if solo_top_n:
            data_start, data_end = self.obtener_rango_top_n(
                ws, first_data, last_data, top_n
            )
        else:
            data_start, data_end = first_data, last_data

        chart = BarChart()
        chart.title = titulo
        chart.y_axis.title = "Subtotal Final"
        chart.x_axis.title = ""

        data_ref = Reference(
            ws,
            min_col=2,
            min_row=data_start - 1,
            max_row=data_end
        )

        cats_ref = Reference(
            ws,
            min_col=1,
            min_row=data_start,
            max_row=data_end
        )

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, "A1")

    def agregar_acumulado_fecha(self, ws, last_data):

        total_row = ws.max_row

        ws.append([
            f"ACUM. AL {ws.cell(row=last_data-1, column=1).value}",
            f"=B{total_row}-B{last_data}",
            #f"=C{total_row}-C{last_data}",
            #f"=D{total_row}-D{last_data}",
            #f"=E{total_row}-E{last_data}",
            #f"=F{total_row}-F{last_data}",
            #""
        ])

        self.formatear_totales(ws, ws.max_row)

    def formatear_header(self, ws, row):
        for cell in ws[row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER_THIN

    def formatear_numeros(self, ws, first_row, last_row):
        for row in ws.iter_rows(
            min_row=first_row,
            max_row=last_row,
            min_col=2,
            max_col=2
        ):
            for cell in row:
                cell.number_format = NUM_FMT
                cell.border = BORDER_THIN

    def formatear_totales(self, ws, row):
        for cell in ws[row]:
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER_THIN
            cell.number_format = NUM_FMT

    def ajustar_columnas(self, ws, ancho_col_a):
        """
        ancho_col_a: ancho específico de la columna A para esta hoja
        """
        widths = {
            "A": ancho_col_a,
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 20,
            "F": 18,
            "G": 12,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def obtener_rango_top_n(self, ws, first_data, last_data, n):

        top_last = min(first_data + n - 1, last_data)

        return first_data, top_last


def index(request, pk=None):
    #files = os.path.join(settings.BASE_DIR, 'ventas', 'ReporteDetallado.xlsx')
    files = os.path.join(settings.BASE_DIR, 'ventas', 'ReporteDetallado.xls')
        
    #df = pd.read_excel(files, decimal='.', thousands=',' )

    df = pd.read_csv(files, encoding='latin1', sep='\t', low_memory=False )

    # Filtrado por Comprobantes
    comprobantes = ['FCVTA', 'DVVTA']
    filtrado = df[df['Comprobante'].isin(comprobantes)].copy()

    # Categorias de Ventas
    condiciones = [
    filtrado['Vendedor'].isin([1,3,4,5,6,11,14,16,17,18]), # Vendedoras
    filtrado['Vendedor'].isin([7,]), # Local
    filtrado['Vendedor'].isin([0,9]) # Mayoristas
    ]
    elecciones = ['Vendedoras', 'Local', 'Mayoristas']

    #print(f'Nro de registros despues del filtrado por comprobantes: {len(filtrado)}')
    filtrado['Canal'] = np.select(condiciones, elecciones, default='Otros')
    
    # Conversion de tipos
    columna_int= ['Vendedor']
    filtrado[columna_int] = filtrado[columna_int].apply(pd.to_numeric, errors='coerce').fillna(-1).astype(int)
    columnas_numericas = ['Subtotal Neto', 'Subtotal Final', 'Precio de compra Neto', 'Bultos Total']
    filtrado[columnas_numericas] = filtrado[columnas_numericas].apply(pd.to_numeric, errors='coerce').fillna(0)
    # Ajuste Subtotal Neto para Informado NO 
    condiciones_a = filtrado['Informado'] == 'NO'

    filtrado['Subtotal Neto'] = np.where(condiciones_a, filtrado['Subtotal Final'], filtrado['Subtotal Neto'])

    # Rellenar Descripcion Vendedor donde falte con 'Jefe'
    cond_vendedor =filtrado['Descripcion Vendedor'].isna() & filtrado['Vendedor'].notna()
    filtrado.loc[cond_vendedor, 'Descripcion Vendedor'] = 'Jefe'

    cond_rechazado = filtrado['Descripcion Motivo Rechazo / Devolucion'].isna() 
    filtrado.loc[cond_rechazado, 'Descripcion Motivo Rechazo / Devolucion'] = ''

    # Rellenamos Documento con Descripcion Comprobante + Serie + Numero
    filtrado['Documento'] = filtrado['Descripcion Comprobante'].astype(str) + ' - ' + filtrado[r'Serie \ Punto de venta'].astype(str) + ' - ' + filtrado['Numero'].astype(str)
    # Agregamos la Columna Costo Neto Facturas
    filtrado['Costo Neto'] = filtrado['Precio de compra Neto'] * 1.21 * filtrado['Bultos Total']
    nueva_lista = ['Comprobante', 'Documento', 'Informado', 'Fecha Comprobante', 'Cliente', 'Razon Social',
                   'Codigo de Articulo', 'Descripcion de Articulo', 'Unidades por Bulto', 
                   'Bultos Total', 'Subtotal Neto', 'Subtotal Final','Motivo Rechazo / Devolucion',
                   'Descripcion Motivo Rechazo / Devolucion', 'Vendedor', 'Descripcion Vendedor', 'Cod. Postal',
                   'Localidad', 'Canal', 'Costo Neto']
    # Dataframe reordenado
    general = filtrado.reindex(columns=nueva_lista)
    general.rename(columns={
        'Fecha Comprobante': 'Fecha_Comprobante',
        'Razon Social': 'Razon_Social',
        'Codigo de Articulo': 'Codigo_Articulo',
        'Descripcion de Articulo': 'Descripcion_Articulo',
        'Unidades por Bulto': 'Unidades_Bulto',
        'Bultos Total': 'Bulto_Total',
        'Subtotal Neto': 'Subtotal_Neto',
        'Motivo Rechazo / Devolucion': 'Motivo_Rechazo_Devolucion',
        'Descripcion Motivo Rechazo / Devolucion': 'Descripcion_Motivo_Rechazo_Devolucion',
        'Descripcion Vendedor': 'Descripcion_Vendedor',
        'Costo Neto': 'Costo_Neto' 
    },inplace=True)


    Canal = general.groupby(['Canal', 'Comprobante']).agg({
        'Subtotal Neto': 'sum',
        'Subtotal Final': 'sum',
        'Costo Neto': 'sum'
    }).reset_index()

    return general.head(350).to_dict(orient='records')
    
    # Agrupacion por Categoria y Comprobantes
    resumen = filtrado.groupby(['Canal', 'Comprobante']).agg({
        'Subtotal Neto':'sum',
        'Subtotal Final':'sum',
        'Costo Neto':'sum'
    }).reset_index()

    Data = []
    for cat in elecciones:
        fcvta = resumen[(resumen['Canal'] == cat) & (resumen['Comprobante'] == 'FCVTA')]
        dvvta = resumen[(resumen['Canal'] == cat) & (resumen['Comprobante'] == 'DVVTA')]
        Data.append( {
            'Canal': cat,
            'SubTotalFcvta': fcvta['Subtotal Neto'].values[0] if not fcvta.empty else 0,
            'TotalIvaFcvta': (fcvta['Subtotal Final'] - fcvta['Subtotal Neto']).values[0] if not fcvta.empty else 0,
            'TotalFcvta': fcvta['Subtotal Final'].values[0] if not fcvta.empty else 0,
            'TotalCostoFcvta': fcvta['Costo Neto'].values[0] if not fcvta.empty else 0,
            'UtilidadFcvta':  0,
            'Cmg_Fcvta': 0,
            'SubTotalDvvta': dvvta['Subtotal Neto'].values[0] if not dvvta.empty else 0,
            'TotalIvaDvvta': (dvvta['Subtotal Final'] - dvvta['Subtotal Neto']).values[0] if not dvvta.empty else 0,
            'TotalDvvta': dvvta['Subtotal Final'].values[0] if not dvvta.empty else 0,
            'TotalCostoDvvta': dvvta['Costo Neto'].values[0] if not dvvta.empty else 0,
            'UtilidadDvvta':  0,
            'Cmg_Dvvta': 0,
            'SubTotalG' : 0,
            'TotalIvaG' :  0,
            'TotalG' :  0,
            'TotalCostoG':  0,
            'UtilidadG' :  0,
            'Cmg_G' : 0,
            } )

    for item in Data:
        item['SubTotalG'] = item['SubTotalFcvta'] + item['SubTotalDvvta']
        item['TotalIvaG'] = item['TotalIvaFcvta'] + item['TotalIvaDvvta']
        item['TotalG'] = item['TotalFcvta'] + item['TotalDvvta']
        item['TotalCostoG'] = item['TotalCostoFcvta'] + item['TotalCostoDvvta']
        item['UtilidadG'] =  item['SubTotalG'] - item['TotalCostoG']
        if item['SubTotalG'] != 0:
            item['Cmg_G'] = (item['UtilidadG'] / item['SubTotalG']) * 100
        item['UtilidadFcvta'] = item['SubTotalFcvta'] - item['TotalCostoFcvta']
        if item['SubTotalFcvta'] != 0:
            item['Cmg_Fcvta'] = (item['UtilidadFcvta'] / item['SubTotalFcvta']) * 100
        item['UtilidadDvvta'] = item['SubTotalDvvta'] - item['TotalCostoDvvta']
        if item['SubTotalDvvta'] != 0:
            item['Cmg_Dvvta'] = (item['UtilidadDvvta'] / item['SubTotalDvvta']) * 100

    
    Ultimo = {
        'Categoria':'Total General',
        'SubTotalG':Data[0]['SubTotalG'] + Data[1]['SubTotalG'] + Data[2]['SubTotalG'],
        'TotalIvaG':Data[0]['TotalIvaG'] + Data[1]['TotalIvaG'] + Data[2]['TotalIvaG'],
        'TotalG':Data[0]['TotalG'] + Data[1]['TotalG'] + Data[2]['TotalG'],
        'TotalCostoG':Data[0]['TotalCostoG'] + Data[1]['TotalCostoG'] + Data[2]['TotalCostoG'],
        'UtilidadG': Data[0]['UtilidadFcvta'] + Data[1]['UtilidadFcvta'] + Data[2]['UtilidadFcvta'] - (Data[0]['UtilidadDvvta'] + Data[1]['UtilidadDvvta'] + Data[2]['UtilidadDvvta']),
        'Cmg_G': 0,
        'SubTotalFcvta':Data[0]['SubTotalFcvta'] + Data[1]['SubTotalFcvta'] + Data[2]['SubTotalFcvta'],
        'TotalIvaFcvta':Data[0]['TotalIvaFcvta'] + Data[1]['TotalIvaFcvta'] + Data[2]['TotalIvaFcvta'],
        'TotalFcvta':Data[0]['TotalFcvta'] + Data[1]['TotalFcvta'] + Data[2]['TotalFcvta'],
        'TotalCostoFcvta':Data[0]['TotalCostoFcvta'] + Data[1]['TotalCostoFcvta'] + Data[2]['TotalCostoFcvta'],
        'UtilidadFcvta':Data[0]['UtilidadFcvta'] + Data[1]['UtilidadFcvta'] + Data[2]['UtilidadFcvta'],
        'Cmg_Fcvta': 0,
        'SubTotalDvvta':Data[0]['SubTotalDvvta'] + Data[1]['SubTotalDvvta'] + Data[2]['SubTotalDvvta'],
        'TotalIvaDvvta':Data[0]['TotalIvaDvvta'] + Data[1]['TotalIvaDvvta'] + Data[2]['TotalIvaDvvta'],
        'TotalDvvta':Data[0]['TotalDvvta'] + Data[1]['TotalDvvta'] + Data[2]['TotalDvvta'],
        'TotalCostoDvvta':Data[0]['TotalCostoDvvta'] + Data[1]['TotalCostoDvvta'] + Data[2]['TotalCostoDvvta'],
        'UtilidadDvvta':Data[0]['UtilidadDvvta'] + Data[1]['UtilidadDvvta'] + Data[2]['UtilidadDvvta'],
        'Cmg_Dvvta': 0,
        } 
    Ultimo['UtilidadG'] = Ultimo['SubTotalG'] - Ultimo['TotalCostoG']
    if Ultimo['SubTotalG'] != 0:
        Ultimo['Cmg_G'] = (Ultimo['UtilidadG'] / Ultimo['SubTotalG']) * 100
    Ultimo['UtilidadFcvta'] = Ultimo['SubTotalFcvta'] - Ultimo['TotalCostoFcvta']
    if Ultimo['SubTotalFcvta'] != 0:
        Ultimo['Cmg_Fcvta'] = (Ultimo['UtilidadFcvta'] / Ultimo['SubTotalFcvta']) * 100
    Ultimo['UtilidadDvvta'] = Ultimo['SubTotalDvvta'] - Ultimo['TotalCostoDvvta']
    if Ultimo['SubTotalDvvta'] != 0:
        Ultimo['Cmg_Dvvta'] = (Ultimo['UtilidadDvvta'] / Ultimo['SubTotalDvvta']) * 100


    # Agrupacion por Vendedor
    VentasXVendedor = []
    Listas_ven = filtrado['Vendedor'].unique().tolist()
    Vendedor = filtrado.groupby(['Vendedor','Comprobante','Descripcion Vendedor','Descripcion Comprobante']).agg({
        'Subtotal Neto':'sum',
        'Subtotal Final':'sum',
        'Costo Neto':'sum'
    }).reset_index()

    for ven in Listas_ven:
        fcvta = Vendedor[(Vendedor['Vendedor'] == ven) & (Vendedor['Comprobante'] == 'FCVTA')]
        dvvta = Vendedor[(Vendedor['Vendedor'] == ven) & (Vendedor['Comprobante'] == 'DVVTA')]
        VentasXVendedor.append( {
            'Vendedor': ven,
            'NombreVendedor': fcvta['Descripcion Vendedor'].values[0] if not fcvta.empty  else 'Jefe de Ventas',
            'SubTotalFcvta': fcvta['Subtotal Neto'].values[0] if not fcvta.empty else 0,
            'TotalIvaFcvta': (fcvta['Subtotal Final'] - fcvta['Subtotal Neto']).values[0] if not fcvta.empty else 0,
            'TotalFcvta': fcvta['Subtotal Final'].values[0] if not fcvta.empty else 0,
            'TotalCostoFcvta': fcvta['Costo Neto'].values[0] if not fcvta.empty else 0,
            'UtilidadFcvta':  0,
            'Cmg_Fcvta': 0,
            'SubTotalDvvta': dvvta['Subtotal Neto'].values[0] if not dvvta.empty else 0,
            'TotalIvaDvvta': (dvvta['Subtotal Final'] - dvvta['Subtotal Neto']).values[0] if not dvvta.empty else 0,
            'TotalDvvta': dvvta['Subtotal Final'].values[0] if not dvvta.empty else 0,
            'TotalCostoDvvta': dvvta['Costo Neto'].values[0] if not dvvta.empty else 0,
            'UtilidadDvvta':  0,
            'Cmg_Dvvta': 0,
            'SubTotalG' : 0,
            'TotalIvaG' : 0,
            'TotalG' : 0,
            'TotalCostoG': 0,
            'UtilidadG' :  0,
            'Cmg_G' : 0,
            } )

    for item in VentasXVendedor:
        item['SubTotalG'] = item['SubTotalFcvta'] + item['SubTotalDvvta']
        item['TotalIvaG'] = item['TotalIvaFcvta'] + item['TotalIvaDvvta']
        item['TotalG'] = item['TotalFcvta'] + item['TotalDvvta']
        item['TotalCostoG'] = item['TotalCostoFcvta'] + item['TotalCostoDvvta']
        item['UtilidadG'] =  item['SubTotalG'] - item['TotalCostoG']
        if item['SubTotalG'] != 0:
            item['Cmg_G'] = (item['UtilidadG'] / item['SubTotalG']) * 100
        item['UtilidadFcvta'] = item['SubTotalFcvta'] - item['TotalCostoFcvta']
        if item['SubTotalFcvta'] != 0:
            item['Cmg_Fcvta'] = (item['UtilidadFcvta'] / item['SubTotalFcvta']) * 100
        item['UtilidadDvvta'] = item['SubTotalDvvta'] - item['TotalCostoDvvta']
        if item['SubTotalDvvta'] != 0:
            item['Cmg_Dvvta'] = (item['UtilidadDvvta'] / item['SubTotalDvvta']) * 100

    # Agrupacion por Localidades
    VentasXLocalidad = []
    Listas_loc = filtrado['Localidad'].unique().tolist()
    Localidad = filtrado.groupby(['Cod. Postal','Localidad','Comprobante']).agg({
        'Subtotal Neto':'sum',
        'Subtotal Final':'sum',
        'Costo Neto':'sum'
    }).reset_index()

    for loc in Listas_loc:
        fcvta = Localidad[(Localidad['Localidad'] == loc) & (Localidad['Comprobante'] == 'FCVTA')]
        dvvta = Localidad[(Localidad['Localidad'] == loc) & (Localidad['Comprobante'] == 'DVVTA')]
        VentasXLocalidad.append( {
            'Localidad': loc,
            'CodPostal': fcvta['Cod. Postal'].values[0] if not fcvta.empty  else (dvvta['Cod. Postal'].values[0] if not dvvta.empty else ''),
            'SubTotalFcvta': fcvta['Subtotal Neto'].values[0] if not fcvta.empty else 0,
            'TotalIvaFcvta': (fcvta['Subtotal Final'] - fcvta['Subtotal Neto']).values[0] if not fcvta.empty else 0,
            'TotalFcvta': fcvta['Subtotal Final'].values[0] if not fcvta.empty else 0,
            'TotalCostoFcvta': fcvta['Costo Neto'].values[0] if not fcvta.empty else 0,
            'UtilidadFcvta':  0,
            'Cmg_Fcvta': 0,
            'SubTotalDvvta': dvvta['Subtotal Neto'].values[0] if not dvvta.empty else 0,
            'TotalIvaDvvta': (dvvta['Subtotal Final'] - dvvta['Subtotal Neto']).values[0] if not dvvta.empty else 0,
            'TotalDvvta': dvvta['Subtotal Final'].values[0] if not dvvta.empty else 0,
            'TotalCostoDvvta': dvvta['Costo Neto'].values[0] if not dvvta.empty else 0,
            'UtilidadDvvta':  0,
            'Cmg_Dvvta': 0,
            'SubTotalG' : 0,
            'TotalIvaG' : 0,
            'TotalG' : 0,
            'TotalCostoG': 0,
            'UtilidadG' :  0,
            'Cmg_G' : 0,
        } )

    for item in VentasXLocalidad:
        item['SubTotalG'] = item['SubTotalFcvta'] + item['SubTotalDvvta']
        item['TotalIvaG'] = item['TotalIvaFcvta'] + item['TotalIvaDvvta']
        item['TotalG'] = item['TotalFcvta'] + item['TotalDvvta']
        item['TotalCostoG'] = item['TotalCostoFcvta'] + item['TotalCostoDvvta']
        item ['UtilidadG'] = item ['SubTotalG'] - item ['TotalCostoG']
        if item ['SubTotalG'] != 0:
            item ['Cmg_G'] = (item ['UtilidadG'] / item ['SubTotalG']) * 100
        item ['UtilidadFcvta'] = item ['SubTotalFcvta'] - item ['TotalCostoFcvta']
        if item ['SubTotalFcvta'] != 0:
            item ['Cmg_Fcvta'] = (item ['UtilidadFcvta'] / item ['SubTotalFcvta']) * 100
        item ['UtilidadDvvta'] = item ['SubTotalDvvta'] - item ['TotalCostoDvvta']
        if item ['SubTotalDvvta'] != 0:
            item ['Cmg_Dvvta'] = (item ['UtilidadDvvta'] / item ['SubTotalDvvta']) * 100
        
    # Agrupacion por Productos

    VentasXProducto = []
    Listas_prod = filtrado['Codigo de Articulo'].unique().tolist()
    Productos = filtrado.groupby(['Codigo de Articulo','Descripcion de Articulo','Comprobante']).agg({
        'Subtotal Neto':'sum',
        'Subtotal Final':'sum',
        'Costo Neto':'sum'
    }).reset_index()

    for prod in Listas_prod:
        fcvta = Productos[(Productos['Codigo de Articulo'] == prod) & (Productos['Comprobante'] == 'FCVTA')]
        dvvta = Productos[(Productos['Codigo de Articulo'] == prod) & (Productos['Comprobante'] == 'DVVTA')]
        VentasXProducto.append( {
            'Producto': prod,
            'Descripcion': fcvta['Descripcion de Articulo'].values[0] if not fcvta.empty  else (dvvta['Descripcion de Articulo'].values[0] if not dvvta.empty else ''),
            'SubTotalFcvta': fcvta['Subtotal Neto'].values[0] if not fcvta.empty else 0,
            'TotalIvaFcvta': (fcvta['Subtotal Final'] - fcvta['Subtotal Neto']).values[0] if not fcvta.empty else 0,
            'TotalFcvta': fcvta['Subtotal Final'].values[0] if not fcvta.empty else 0,
            'TotalCostoFcvta': fcvta['Costo Neto'].values[0] if not fcvta.empty else 0,
            'UtilidadFcvta':  0,
            'Cmg_Fcvta': 0,
            'SubTotalDvvta': dvvta['Subtotal Neto'].values[0] if not dvvta.empty else 0,
            'TotalIvaDvvta': (dvvta['Subtotal Final'] - dvvta['Subtotal Neto']).values[0] if not dvvta.empty else 0,
            'TotalDvvta': dvvta['Subtotal Final'].values[0] if not dvvta.empty else 0,
            'TotalCostoDvvta': dvvta['Costo Neto'].values[0] if not dvvta.empty else 0,
            'UtilidadDvvta':  0,
            'Cmg_Dvvta': 0,
            'SubTotalG' : 0,
            'TotalIvaG' : 0,
            'TotalG' : 0,
            'TotalCostoG': 0,
            'UtilidadG' :  0,
            'Cmg_G' : 0
        } )

    for item in VentasXProducto:
        item ['SubTotalG'] = item ['SubTotalFcvta'] + item ['SubTotalDvvta']
        item ['TotalIvaG'] = item ['TotalIvaFcvta'] + item ['TotalIvaDvvta']
        item ['TotalG'] = item ['TotalFcvta'] + item ['TotalDvvta']
        item ['TotalCostoG'] = item ['TotalCostoFcvta'] + item ['TotalCostoDvvta']
        item ['UtilidadG'] = item ['SubTotalG'] - item ['TotalCostoG']
        if item ['SubTotalG'] != 0:
            item ['Cmg_G'] = (item ['UtilidadG'] / item ['SubTotalG']) * 100
        item ['UtilidadFcvta'] = item ['SubTotalFcvta'] - item ['TotalCostoFcvta']
        if item ['SubTotalFcvta'] != 0:
            item ['Cmg_Fcvta'] = (item ['UtilidadFcvta'] / item ['SubTotalFcvta']) * 100
        item ['UtilidadDvvta'] = item ['SubTotalDvvta'] - item ['TotalCostoDvvta']
        if item ['SubTotalDvvta'] != 0:
            item ['Cmg_Dvvta'] = (item ['UtilidadDvvta'] / item ['SubTotalDvvta']) * 100


    return render(request, 'ventas.html', {
        'Data': Data,
        'VentasXVendedor': VentasXVendedor,'ult': Ultimo,
        'VentasXLocalidad': VentasXLocalidad,
        'VentasXProducto': VentasXProducto})